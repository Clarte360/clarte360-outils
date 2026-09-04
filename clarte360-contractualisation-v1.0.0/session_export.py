from __future__ import annotations

import io
import json
import zipfile
from datetime import date, datetime, time
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook

from xlsm_safe import workbook_values

CLEAR_TOKEN = "__CLEAR__"

CANONICAL_FIN_HEADERS = [
    'ID_FINANCEMENT','NO_CLAR','ORDRE','ORDRE_FINANCEMENT','TYPE_FINANCEUR','NOM_FINANCEUR',
    'SIRET_FINANCEUR','ADRESSE_FINANCEUR','CODE_POST_FINANCEUR','CP_FINANCEUR','VILLE_FINANCEUR',
    'CONTACT_FINANCEUR','EMAIL_FACTURATION','EMAIL_FINANCEUR','MONTANT_HT','TAUX_TVA','MONTANT_TVA',
    'MONTANT_TTC','FACTURE_A_ETABLIR_A','REFERENCE_PRISE_EN_CHARGE','STATUT_FINANCEMENT','DATE_ACCORD','OBSERVATIONS'
]


def _json_default(value: Any):
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    raise TypeError(f'Non serialisable: {type(value)!r}')


def workbook_headers(data: bytes, sheet_name: str) -> List[str]:
    wb = workbook_values(data, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        return [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1) if str(ws.cell(1, c).value or '').strip()]
    finally:
        wb.close()


def next_available_action(data: bytes, reserved: Iterable[str] = ()) -> Tuple[int, str]:
    reserved_set = {str(x).strip().upper() for x in reserved}
    wb = workbook_values(data, data_only=True)
    try:
        ws = wb['CONV ADM']
        headers = {str(ws.cell(1, c).value or '').strip(): c for c in range(1, ws.max_column + 1)}
        no_col = headers.get('NO_CLAR')
        if not no_col:
            raise RuntimeError('Colonne NO_CLAR introuvable dans CONV ADM')
        check_cols = [headers.get(x) for x in ['NOM_ENT','INTITULE_FORMA','PRENOM_STAGIAIRE','EMAIL','Date_debut_action']]
        for r in range(2, ws.max_row + 1):
            no = str(ws.cell(r, no_col).value or '').strip().upper()
            if not no or no in reserved_set:
                continue
            if all(ws.cell(r, c).value in (None, '') for c in check_cols if c):
                return r, no
    finally:
        wb.close()
    raise RuntimeError('Aucune ligne libre disponible dans CONV ADM')


def read_financements_for_action(data: bytes, no_clar: str) -> List[Dict[str, Any]]:
    wb = workbook_values(data, data_only=True)
    try:
        if 'FINANCEMENTS' not in wb.sheetnames:
            return []
        ws = wb['FINANCEMENTS']
        headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        if 'NO_CLAR' not in headers:
            return []
        no_idx = headers.index('NO_CLAR') + 1
        rows = []
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, no_idx).value or '').strip().upper() != no_clar.strip().upper():
                continue
            row = {}
            for c, h in enumerate(headers, start=1):
                if h:
                    row[h] = ws.cell(r, c).value
            rows.append(row)
        return rows
    finally:
        wb.close()


def _write_value(cell, value: Any, header: str):
    if value is None:
        return
    cell.value = value
    if isinstance(value, (datetime, date)) or header.upper().startswith('DATE') or header in {'Date_debut_action','Date_de_fin_d_action'}:
        if isinstance(value, str):
            # Keep explicit control tokens/text as text; actual ISO dates are converted below.
            try:
                value_date = datetime.fromisoformat(value[:10]).date()
                cell.value = value_date
            except Exception:
                pass
        if isinstance(cell.value, (datetime, date)):
            cell.number_format = 'dd/mm/yyyy'
    elif isinstance(value, time):
        cell.number_format = 'hh:mm'
    elif isinstance(value, float) and ('MONTANT' in header or header in {'INTRA_HT','TTC','TVA'}):
        cell.number_format = '#,##0.00 [$€-fr-FR]'


def build_import_xlsx(records: List[Dict[str, Any]], conv_headers: List[str], fin_headers: List[str]) -> bytes:
    if not conv_headers:
        raise RuntimeError('Entêtes CONV ADM introuvables')
    if not fin_headers:
        fin_headers = CANONICAL_FIN_HEADERS

    wb = Workbook()
    ws = wb.active
    ws.title = 'CONV ADM'
    ws.append(conv_headers)
    for rec in records:
        values = dict(rec.get('conv_values') or {})
        values['NO_CLAR'] = rec['no_clar']
        row = []
        for h in conv_headers:
            row.append(values.get(h, None))
        ws.append(row)
        rr = ws.max_row
        for c, h in enumerate(conv_headers, start=1):
            _write_value(ws.cell(rr, c), values.get(h, None), h)

    wf = wb.create_sheet('FINANCEMENTS')
    wf.append(fin_headers)
    for rec in records:
        for frow in rec.get('financements') or []:
            row = [frow.get(h, None) for h in fin_headers]
            wf.append(row)
            rr = wf.max_row
            for c, h in enumerate(fin_headers, start=1):
                _write_value(wf.cell(rr, c), frow.get(h, None), h)

    wm = wb.create_sheet('META')
    wm.append(['NO_CLAR','OPERATION','TYPE_PRESTATION','TYPE_CONTRACTUALISATION','SOURCE','MODELE_CONTRAT','DATE_PREPARATION'])
    for rec in records:
        wm.append([
            rec['no_clar'], rec.get('operation','MISE_A_JOUR'), rec.get('prestation',''), rec.get('contract_type',''),
            rec.get('source',''), rec.get('model_version',''), datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        ])

    for sh in (ws, wf, wm):
        sh.freeze_panes = 'A2'
        sh.auto_filter.ref = sh.dimensions
        for cell in sh[1]:
            cell.font = cell.font.copy(bold=True)
        # pragmatic widths
        for col in sh.columns:
            letter = col[0].column_letter
            maxlen = min(60, max(10, max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2))
            sh.column_dimensions[letter].width = maxlen

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_session_json(records: List[Dict[str, Any]]) -> bytes:
    clean = []
    for rec in records:
        clean.append({k: v for k, v in rec.items() if k not in {'pdf_bytes','contract_json_bytes'}})
    payload = {
        'format': 'CLARTE360_IMPORT_SESSION_V1',
        'created_at': datetime.now().isoformat(timespec='seconds'),
        'actions': clean,
    }
    return json.dumps(payload, ensure_ascii=False, indent=2, default=_json_default).encode('utf-8')


def build_session_zip(records: List[Dict[str, Any]], import_xlsx: bytes, session_json: bytes) -> bytes:
    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as z:
        z.writestr('CLARTE360_IMPORT_SESSION.xlsx', import_xlsx)
        z.writestr('CLARTE360_IMPORT_SESSION.json', session_json)
        for rec in records:
            no = rec['no_clar']
            if rec.get('contract_json_bytes'):
                z.writestr(f'{no}/{no}_dossier_contractuel.json', rec['contract_json_bytes'])
            if rec.get('pdf_bytes'):
                z.writestr(f"{no}/{rec.get('pdf_name') or (no + '_contrat.pdf')}", rec['pdf_bytes'])
    return out.getvalue()
