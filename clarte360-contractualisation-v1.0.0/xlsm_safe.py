from __future__ import annotations

import io
import re
import zipfile
from copy import deepcopy
from datetime import date, datetime, time
from typing import Dict, Iterable, List, Tuple, Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL_DOC = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_REL_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"
ET.register_namespace('', NS_MAIN)
ET.register_namespace('r', NS_REL_DOC)


def _q(tag: str) -> str:
    return f"{{{NS_MAIN}}}{tag}"


def _col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _excel_serial(d: Any) -> float:
    if isinstance(d, datetime):
        dt = d
    elif isinstance(d, date):
        dt = datetime.combine(d, time())
    else:
        raise TypeError(d)
    epoch = datetime(1899, 12, 30)
    return (dt - epoch).total_seconds() / 86400.0


def inspect_workbook(data: bytes) -> Dict[str, Any]:
    with zipfile.ZipFile(io.BytesIO(data), 'r') as z:
        names = set(z.namelist())
        return {
            'has_vba': 'xl/vbaProject.bin' in names,
            'has_custom_ui': any(n.startswith('customUI/') for n in names),
            'zip_entries': len(names),
        }


def workbook_values(data: bytes, data_only: bool = False):
    return load_workbook(io.BytesIO(data), keep_vba=True, keep_links=True, data_only=data_only, read_only=True)


def _sheet_map(z: zipfile.ZipFile) -> Dict[str, str]:
    wb = ET.fromstring(z.read('xl/workbook.xml'))
    rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
    relmap = {}
    for rel in rels:
        rid = rel.attrib.get('Id')
        target = rel.attrib.get('Target')
        if rid and target:
            if target.startswith('/'):
                target = target.lstrip('/')
            elif not target.startswith('xl/'):
                target = 'xl/' + target.lstrip('/')
            relmap[rid] = target
    out = {}
    sheets = wb.find(_q('sheets'))
    if sheets is not None:
        for sh in sheets:
            name = sh.attrib['name']
            rid = sh.attrib.get(f'{{{NS_REL_DOC}}}id')
            if rid in relmap:
                out[name] = relmap[rid]
    return out


def _find_or_create_cell(row_el: ET.Element, ref: str, style_id: str | None = None) -> ET.Element:
    for c in row_el.findall(_q('c')):
        if c.attrib.get('r') == ref:
            return c
    c = ET.Element(_q('c'), {'r': ref})
    if style_id:
        c.attrib['s'] = style_id
    # insert sorted by column index
    def colnum(cellref: str):
        letters = re.match(r'([A-Z]+)', cellref).group(1)
        n = 0
        for ch in letters:
            n = n * 26 + ord(ch) - 64
        return n
    target = colnum(ref)
    cells = list(row_el.findall(_q('c')))
    inserted = False
    for i, existing in enumerate(cells):
        if colnum(existing.attrib['r']) > target:
            row_el.insert(i, c)
            inserted = True
            break
    if not inserted:
        row_el.append(c)
    return c


def _set_cell_value(c: ET.Element, value: Any):
    # Preserve style and reference only; replace formula/value payload.
    for child in list(c):
        c.remove(child)
    if value is None or value == "":
        c.attrib.pop('t', None)
        return
    if isinstance(value, bool):
        c.attrib['t'] = 'b'
        v = ET.SubElement(c, _q('v'))
        v.text = '1' if value else '0'
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        c.attrib.pop('t', None)
        v = ET.SubElement(c, _q('v'))
        v.text = str(value)
    elif isinstance(value, (datetime, date)):
        c.attrib.pop('t', None)
        v = ET.SubElement(c, _q('v'))
        v.text = str(_excel_serial(value))
    elif isinstance(value, time):
        c.attrib.pop('t', None)
        secs = value.hour*3600 + value.minute*60 + value.second
        v = ET.SubElement(c, _q('v'))
        v.text = str(secs / 86400.0)
    else:
        c.attrib['t'] = 'inlineStr'
        is_el = ET.SubElement(c, _q('is'))
        t = ET.SubElement(is_el, _q('t'))
        s = str(value)
        if s.startswith(' ') or s.endswith(' ') or '\n' in s:
            t.attrib['{http://www.w3.org/XML/1998/namespace}space'] = 'preserve'
        t.text = s


def _style_for_col(sheet_root: ET.Element, col_letter: str, preferred_row: int) -> str | None:
    sheetdata = sheet_root.find(_q('sheetData'))
    if sheetdata is None:
        return None
    for rnum in [preferred_row, max(1, preferred_row-1), 2, 3, 1]:
        row = next((r for r in sheetdata.findall(_q('row')) if int(r.attrib.get('r','0')) == rnum), None)
        if row is None:
            continue
        ref = f'{col_letter}{rnum}'
        for c in row.findall(_q('c')):
            if c.attrib.get('r') == ref and c.attrib.get('s'):
                return c.attrib.get('s')
    return None


def patch_conv_adm(data: bytes, row_number: int, values_by_header: Dict[str, Any]) -> bytes:
    src = io.BytesIO(data)
    out = io.BytesIO()
    with zipfile.ZipFile(src, 'r') as zin:
        smap = _sheet_map(zin)
        target = smap['CONV ADM']
        sheet_root = ET.fromstring(zin.read(target))
        sheetdata = sheet_root.find(_q('sheetData'))
        if sheetdata is None:
            raise RuntimeError('sheetData introuvable')
        # headers by parsing row 1, including shared/inline strings via openpyxl for reliability
        wb = workbook_values(data, data_only=False)
        ws = wb['CONV ADM']
        headers = {str(ws.cell(1,c).value): c for c in range(1, ws.max_column+1) if ws.cell(1,c).value}
        wb.close()
        row_el = next((r for r in sheetdata.findall(_q('row')) if int(r.attrib.get('r','0')) == row_number), None)
        if row_el is None:
            row_el = ET.Element(_q('row'), {'r': str(row_number)})
            rows = list(sheetdata.findall(_q('row')))
            inserted=False
            for i,r in enumerate(rows):
                if int(r.attrib.get('r','0')) > row_number:
                    sheetdata.insert(i,row_el); inserted=True; break
            if not inserted:
                sheetdata.append(row_el)
        for header, value in values_by_header.items():
            if header not in headers:
                continue
            col = headers[header]
            letter = _col_letter(col)
            ref = f'{letter}{row_number}'
            style = _style_for_col(sheet_root, letter, row_number)
            c = _find_or_create_cell(row_el, ref, style)
            # do not overwrite formulas unless explicit value supplied and header is not formula-driven
            _set_cell_value(c, value)
        new_sheet = ET.tostring(sheet_root, encoding='utf-8', xml_declaration=True)
        with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                payload = new_sheet if item.filename == target else zin.read(item.filename)
                zout.writestr(item, payload)
    return out.getvalue()


def _read_financing_rows(data: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = workbook_values(data, data_only=False)
    ws = wb['FINANCEMENTS']
    headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    rows=[]
    for r in range(2, ws.max_row+1):
        d={headers[c-1]: ws.cell(r,c).value for c in range(1, len(headers)+1)}
        if any(v not in (None,'') for v in d.values()):
            rows.append(d)
    wb.close()
    return headers, rows


def replace_financements_for_action(data: bytes, no_clar: str, new_rows: List[Dict[str, Any]]) -> bytes:
    headers, existing = _read_financing_rows(data)
    kept = [r for r in existing if str(r.get('NO_CLAR') or '').strip().upper() != no_clar.strip().upper()]
    combined = kept + new_rows
    src = io.BytesIO(data)
    out = io.BytesIO()
    with zipfile.ZipFile(src, 'r') as zin:
        smap = _sheet_map(zin)
        target = smap['FINANCEMENTS']
        root = ET.fromstring(zin.read(target))
        sheetdata = root.find(_q('sheetData'))
        if sheetdata is None:
            raise RuntimeError('sheetData FINANCEMENTS introuvable')
        header_row = next((r for r in sheetdata.findall(_q('row')) if int(r.attrib.get('r','0')) == 1), None)
        template_row = next((r for r in sheetdata.findall(_q('row')) if int(r.attrib.get('r','0')) == 2), None)
        if header_row is None:
            raise RuntimeError('Entête FINANCEMENTS introuvable')
        # Remove all rows except header
        for r in list(sheetdata.findall(_q('row'))):
            if int(r.attrib.get('r','0')) != 1:
                sheetdata.remove(r)
        # At least one blank data row to keep table valid if empty
        data_rows = combined if combined else [{}]
        for idx, d in enumerate(data_rows, start=2):
            r = ET.Element(_q('row'), {'r':str(idx)})
            for ci, header in enumerate(headers, start=1):
                letter = _col_letter(ci)
                ref = f'{letter}{idx}'
                style=None
                if template_row is not None:
                    tc = next((c for c in template_row.findall(_q('c')) if c.attrib.get('r') == f'{letter}2'), None)
                    if tc is not None:
                        style=tc.attrib.get('s')
                c = ET.SubElement(r, _q('c'), {'r':ref})
                if style: c.attrib['s']=style
                _set_cell_value(c, d.get(header))
            sheetdata.append(r)
        # update dimension if present
        dim = root.find(_q('dimension'))
        endrow = max(2, 1+len(data_rows))
        if dim is not None:
            dim.attrib['ref'] = f'A1:S{endrow}'
        new_sheet = ET.tostring(root, encoding='utf-8', xml_declaration=True)

        # find table target from sheet rels
        rel_path = target.replace('worksheets/', 'worksheets/_rels/') + '.rels'
        table_target = None
        if rel_path in zin.namelist():
            relroot=ET.fromstring(zin.read(rel_path))
            for rel in relroot:
                tgt=rel.attrib.get('Target','')
                if 'tables/' in tgt:
                    base='xl/worksheets/'
                    # sheet rel target is usually ../tables/tableN.xml
                    parts=target.split('/')[:-1]
                    import posixpath
                    table_target=posixpath.normpath(posixpath.join('/'.join(parts), tgt))
                    break
        table_xml=None
        if table_target and table_target in zin.namelist():
            tr=ET.fromstring(zin.read(table_target))
            tr.attrib['ref']=f'A1:S{endrow}'
            af=tr.find(_q('autoFilter'))
            if af is not None: af.attrib['ref']=f'A1:S{endrow}'
            table_xml=ET.tostring(tr, encoding='utf-8', xml_declaration=True)

        with zipfile.ZipFile(out,'w',zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename==target:
                    payload=new_sheet
                elif table_xml is not None and item.filename==table_target:
                    payload=table_xml
                else:
                    payload=zin.read(item.filename)
                zout.writestr(item,payload)
    return out.getvalue()


def first_available_conv_row(data: bytes) -> Tuple[int, str]:
    wb = workbook_values(data, data_only=True)
    ws=wb['CONV ADM']
    headers={ws.cell(1,c).value:c for c in range(1,ws.max_column+1)}
    check_cols=[headers.get(x) for x in ['NOM_ENT','INTITULE_FORMA','PRENOM_STAGIAIRE','EMAIL','Date_debut_action']]
    for r in range(2, ws.max_row+1):
        no=ws.cell(r,headers['NO_CLAR']).value
        if not no: continue
        if all(ws.cell(r,c).value in (None,'') for c in check_cols if c):
            wb.close(); return r,str(no)
    wb.close()
    raise RuntimeError('Aucune ligne libre disponible dans CONV ADM')


def action_row(data: bytes, no_clar: str) -> Tuple[int, Dict[str, Any]] | None:
    wb_formula=workbook_values(data, data_only=False)
    wb_values=workbook_values(data, data_only=True)
    wsf=wb_formula['CONV ADM']; wsv=wb_values['CONV ADM']
    headers=[wsf.cell(1,c).value for c in range(1,wsf.max_column+1)]
    for r in range(2,wsf.max_row+1):
        if str(wsf.cell(r,1).value or '').strip().upper()==no_clar.strip().upper():
            d={}
            for c,h in enumerate(headers, start=1):
                if h:
                    vv=wsv.cell(r,c).value
                    d[h]=vv if vv is not None else wsf.cell(r,c).value
            wb_formula.close(); wb_values.close(); return r,d
    wb_formula.close(); wb_values.close(); return None

def force_recalc_on_open(data: bytes) -> bytes:
    src=io.BytesIO(data); out=io.BytesIO()
    with zipfile.ZipFile(src,'r') as zin:
        root=ET.fromstring(zin.read('xl/workbook.xml'))
        calc=root.find(_q('calcPr'))
        if calc is None:
            calc=ET.SubElement(root,_q('calcPr'))
        calc.attrib['calcMode']='auto'
        calc.attrib['fullCalcOnLoad']='1'
        calc.attrib['forceFullCalc']='1'
        xml=ET.tostring(root,encoding='utf-8',xml_declaration=True)
        with zipfile.ZipFile(out,'w',zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                zout.writestr(item, xml if item.filename=='xl/workbook.xml' else zin.read(item.filename))
    return out.getvalue()
