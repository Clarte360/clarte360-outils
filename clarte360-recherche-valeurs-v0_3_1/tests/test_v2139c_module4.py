from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
TEXT=ROOT.joinpath('app.py').read_text(encoding='utf-8')


def test_version_9c():
    assert 'APP_VERSION = "2.2.0-preproduction-4"' in TEXT


def test_clean_referential_is_embedded():
    wb=load_workbook(ROOT/'data'/'referentiel_rvc360.xlsx', read_only=True, data_only=True)
    assert 'Référentiel nettoyé' in wb.sheetnames
    ws=wb['Référentiel nettoyé']
    assert ws['A1'].value == 'Code'
    assert ws['B1'].value == 'Valeur'
    assert ws['D1'].value == 'Définition Clarté360'
    assert ws.max_row == 205


def test_loader_supports_clean_definition_header():
    assert '"Définition Clarté360":"definition"' in TEXT


def test_refusal_does_not_end_cycle_immediately():
    assert 'reorientation_apres_refus' in TEXT
    assert 'Le refus d’une hypothèse ne signifie pas que la situation est épuisée' in TEXT
    assert 'Approfondir un autre aspect de cette situation' in TEXT


def test_reorientation_is_bounded():
    assert 'reorientation_count' in TEXT
    assert 'int(cycle.get("reorientation_count",0))<2' in TEXT


def test_candidate_engine_receives_full_context_and_refusals():
    assert '"contexte_beneficiaire":_module4_context_payload()' in TEXT
    assert '"candidats_deja_refuses":cycle.get("candidate_round_history",[])' in TEXT


def test_one_retained_hypothesis_per_cycle_remains():
    block=TEXT[TEXT.index('def _module4_add_hypothesis'):TEXT.index('def _module4_threshold_invitation')]
    assert 'hypothesis_basket.append(item)' in block
    assert 'cycle["stage"]="termine"' in block
